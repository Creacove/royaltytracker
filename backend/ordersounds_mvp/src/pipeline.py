"""
OrderSounds Main Pipeline
End-to-end processing: PDF → OCR → Normalize → Validate → Export
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional
import uuid

# Add src to path
sys.path.insert(0, str(Path(__file__).parent))

from ocr_processor import OCRProcessor, generate_table_fingerprint
from table_reconstructor import TableReconstructor, merge_bounding_boxes, calculate_row_confidence
from normalizer import RoyaltyNormalizer
from validator import FinancialValidator, add_validation_metadata
from db_writer import DatabaseWriter

import pandas as pd
from dotenv import load_dotenv


class OrderSoundsPipeline:
    """
    Main orchestrator for royalty report processing
    
    Workflow:
    1. OCR: PDF → JSON (Google Document AI)
    2. Reconstruct: JSON → DataFrame (table merging)
    3. Normalize: Clean data (ISRC, territory, platform, currency)
    4. Validate: Check math (99% accuracy target)
    5. Export: Save to database + Excel
    """
    
    def __init__(self, config: Optional[Dict] = None):
        # Load environment config
        load_dotenv()

        default_config = {
            'project_id': os.getenv('GOOGLE_CLOUD_PROJECT'),
            'location': os.getenv('DOCUMENTAI_LOCATION', 'us'),
            'processor_id': os.getenv('DOCUMENTAI_PROCESSOR_ID'),
            'output_dir': Path(os.getenv('OUTPUT_DIR', './outputs')),
            'validation_tolerance': float(os.getenv('VALIDATION_TOLERANCE', 0.01)),
            'database_url': os.getenv('DATABASE_URL'),
            'enable_db': os.getenv('ENABLE_DB', 'true').lower() in ['1', 'true', 'yes'],
            'credentials_path': os.getenv('GOOGLE_APPLICATION_CREDENTIALS'),
        }

        if config:
            merged = default_config.copy()
            merged.update(config)
            self.config = merged
        else:
            self.config = default_config
        
        # Create output directories
        self.output_dir = Path(self.config['output_dir'])
        self.ocr_dir = self.output_dir / 'ocr'
        self.data_dir = self.output_dir / 'data'
        self.reports_dir = self.output_dir / 'reports'
        
        for dir_path in [self.ocr_dir, self.data_dir, self.reports_dir]:
            dir_path.mkdir(parents=True, exist_ok=True)
        
        # Validate required config
        self._validate_config()

        # Initialize processors
        self.ocr_processor = OCRProcessor(
            project_id=self.config['project_id'],
            location=self.config['location'],
            processor_id=self.config['processor_id'],
            credentials_path=self.config.get('credentials_path')
        )
        
        self.table_reconstructor = TableReconstructor()
        self.normalizer = RoyaltyNormalizer()
        self.validator = FinancialValidator(tolerance=self.config['validation_tolerance'])

        # Optional DB writer
        self.db_writer = None
        if self.config.get('enable_db') and self.config.get('database_url'):
            self.db_writer = DatabaseWriter(self.config['database_url'])
    
    def process_cmo_report(
        self,
        pdf_path: str,
        cmo_name: Optional[str] = None,
        report_metadata: Optional[Dict] = None,
        write_db: Optional[bool] = None
    ) -> Dict:
        """
        Process complete CMO report end-to-end
        
        Args:
            pdf_path: Path to PDF file
            cmo_name: Name of CMO (e.g., "COSON Nigeria")
            report_metadata: Additional metadata (period_start, period_end, etc.)
            
        Returns:
            Processing result with paths to outputs
        """
        print(f"\n{'='*70}")
        print(f"OrderSounds Pipeline - Processing CMO Report")
        print(f"{'='*70}\n")
        print(f"PDF: {pdf_path}")
        print(f"CMO: {cmo_name or 'Unknown'}\n")
        
        start_time = datetime.now()
        document_id = str(uuid.uuid4())

        if not Path(pdf_path).exists():
            return {'status': 'failed', 'error': f'PDF not found: {pdf_path}'}
        
        # Stage 1: OCR
        print("STAGE 1: OCR Processing")
        print("-" * 70)
        ocr_output = self.ocr_processor.process_pdf(pdf_path, self.ocr_dir)
        ocr_json_path = self.ocr_dir / f"{Path(pdf_path).stem}_ocr.json"
        
        # Generate table fingerprint
        table_headers = None
        if ocr_output['all_tables']:
            first_table = ocr_output['all_tables'][0]
            table_headers = [h['text'] for h in first_table['headers']]
            fingerprint = generate_table_fingerprint(table_headers)
        else:
            fingerprint = None
        
        # Stage 2: Table Reconstruction
        print("\nSTAGE 2: Table Reconstruction")
        print("-" * 70)
        df_raw = self.table_reconstructor.reconstruct_from_ocr(ocr_output)
        
        if df_raw.empty:
            print("❌ No data extracted. Check OCR output.")
            return {'status': 'failed', 'error': 'No data extracted'}
        
        # Add evidence linking
        df_raw = merge_bounding_boxes(df_raw)
        df_raw = calculate_row_confidence(df_raw)

        # Align page/row fields with DB schema
        df_raw = df_raw.rename(columns={
            '_page_number': 'page_number',
            '_row_position': 'row_position'
        })
        
        # Stage 3: Normalization
        print("\nSTAGE 3: Data Normalization")
        print("-" * 70)
        df_normalized = self.normalizer.normalize_dataframe(df_raw)
        
        # Stage 4: Validation
        print("\nSTAGE 4: Financial Validation")
        print("-" * 70)
        
        # Add transaction IDs
        df_normalized['transaction_id'] = [str(uuid.uuid4()) for _ in range(len(df_normalized))]
        df_normalized['document_id'] = document_id
        
        validation_result = self.validator.validate_dataframe(df_normalized)
        
        # Add validation metadata to DataFrame
        df_final = add_validation_metadata(df_normalized, validation_result)
        
        # Stage 5: Export
        print("\nSTAGE 5: Export")
        print("-" * 70)
        
        # Export to CSV (for database import)
        csv_path = self.data_dir / f"{Path(pdf_path).stem}_normalized.csv"
        df_final.to_csv(csv_path, index=False)
        print(f"✓ CSV: {csv_path}")
        
        # Export to Excel (for human review)
        excel_path = self.reports_dir / f"{Path(pdf_path).stem}_report.xlsx"
        self._export_to_excel(df_final, excel_path, validation_result, cmo_name)
        print(f"✓ Excel: {excel_path}")
        
        # Export validation errors
        if validation_result['errors']:
            errors_path = self.reports_dir / f"{Path(pdf_path).stem}_errors.json"
            self.validator.export_error_report(str(errors_path))
        
        # Export processing metadata
        processing_time = (datetime.now() - start_time).total_seconds()
        
        metadata = {
            'document_id': document_id,
            'pdf_path': pdf_path,
            'original_filename': Path(pdf_path).name,
            'cmo_name': cmo_name,
            'table_fingerprint': fingerprint,
            'processing_timestamp': datetime.now().isoformat(),
            'processing_time_seconds': processing_time,
            'total_pages': ocr_output['metadata']['total_pages'],
            'total_transactions': len(df_final),
            'validation': {
                'accuracy_score': validation_result['accuracy_score'],
                'passed': validation_result['passed'],
                'critical_errors': validation_result['critical_errors'],
                'warning_errors': validation_result['warning_errors'],
            },
            'ocr_confidence_avg': df_final['ocr_confidence'].mean(),
            'outputs': {
                'ocr_json': str(ocr_json_path),
                'normalized_csv': str(csv_path),
                'excel_report': str(excel_path),
            }
        }
        
        if report_metadata:
            metadata.update(report_metadata)
        
        metadata_path = self.reports_dir / f"{Path(pdf_path).stem}_metadata.json"
        with open(metadata_path, 'w') as f:
            json.dump(metadata, f, indent=2)
        print(f"✓ Metadata: {metadata_path}")
        
        # Optional: Write to database
        db_status = None
        if write_db is None:
            write_db = bool(self.db_writer)

        if write_db and self.db_writer:
            print("\nSTAGE 6: Database Ingestion")
            print("-" * 70)
            try:
                db_status = self.db_writer.write_report(
                    df_final,
                    document_id=document_id,
                    pdf_path=pdf_path,
                    ocr_json_path=str(ocr_json_path),
                    cmo_name=cmo_name,
                    table_fingerprint=fingerprint,
                    table_headers=table_headers,
                    validation_result=validation_result,
                    processing_time_seconds=processing_time,
                    report_metadata={
                        'original_filename': Path(pdf_path).name,
                        'report_period_start': report_metadata.get('report_period_start') if report_metadata else None,
                        'report_period_end': report_metadata.get('report_period_end') if report_metadata else None,
                        'total_pages': ocr_output['metadata']['total_pages'],
                        'ocr_confidence_avg': df_final['ocr_confidence'].mean(),
                    }
                )
                print(f"âœ… DB ingestion complete: {db_status.get('transactions_inserted')} transactions")
            except Exception as exc:
                db_status = {'status': 'failed', 'error': str(exc)}
                print(f"âŒ DB ingestion failed: {exc}")

        if db_status:
            metadata['db'] = db_status
            with open(metadata_path, 'w') as f:
                json.dump(metadata, f, indent=2)

        # Summary
        print(f"\n{'='*70}")
        print(f"PROCESSING COMPLETE")
        print(f"{'='*70}")
        print(f"Time:              {processing_time:.1f}s")
        print(f"Transactions:      {len(df_final):,}")
        print(f"Accuracy:          {validation_result['accuracy_score']:.2f}%")
        print(f"Status:            {'✅ PASSED' if validation_result['passed'] else '❌ FAILED'}")
        print(f"{'='*70}\n")
        
        return metadata

    def _validate_config(self):
        missing = []
        for key in ['project_id', 'processor_id', 'location']:
            if not self.config.get(key):
                missing.append(key)
        if missing:
            raise ValueError(
                f"Missing required config values: {missing}. "
                f"Check your .env or pass config explicitly."
            )
    
    def _export_to_excel(
        self,
        df: pd.DataFrame,
        excel_path: Path,
        validation_result: Dict,
        cmo_name: Optional[str]
    ):
        """Export to Excel with formatting and summary sheet"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["OrderSounds Royalty Report"],
            [""],
            ["CMO Name", cmo_name or "Unknown"],
            ["Report Date", datetime.now().strftime("%Y-%m-%d")],
            [""],
            ["STATISTICS"],
            ["Total Transactions", len(df)],
            ["Valid Transactions", validation_result['valid_rows']],
            ["Critical Errors", validation_result['critical_errors']],
            ["Warning Errors", validation_result['warning_errors']],
            ["Accuracy Score", f"{validation_result['accuracy_score']:.2f}%"],
            [""],
            ["FINANCIAL SUMMARY"],
            ["Total Gross Revenue", f"${df['gross_revenue'].sum():,.2f}"],
            ["Total Net Revenue", f"${df['net_revenue'].sum():,.2f}"],
            ["Total Publisher Share", f"${df['publisher_share'].sum():,.2f}"],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Format summary sheet
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
        
        # Sheet 2: Transactions
        ws_trans = wb.create_sheet("Transactions")
        
        # Select columns to export (exclude internal metadata)
        export_cols = [
            'transaction_id', 'isrc', 'track_title', 'track_artist',
            'platform', 'territory', 'usage_count',
            'gross_revenue', 'commission', 'net_revenue', 'publisher_share',
            'sales_start', 'sales_end',
            'validation_passed', 'ocr_confidence'
        ]
        
        df_export = df[[col for col in export_cols if col in df.columns]]
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
            ws_trans.append(row)
            
            # Highlight failed validations
            if r_idx > 1 and not df_export.iloc[r_idx - 2]['validation_passed']:
                for col_idx in range(1, len(row) + 1):
                    ws_trans.cell(r_idx, col_idx).fill = PatternFill(
                        start_color='FFE6E6', fill_type='solid'
                    )
        
        # Format header row
        for cell in ws_trans[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Auto-adjust column widths
        for column in ws_trans.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_trans.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save
        wb.save(excel_path)


def main():
    """Command-line interface"""
    import argparse
    
    parser = argparse.ArgumentParser(description='OrderSounds Royalty Report Processor')
    parser.add_argument('pdf_path', help='Path to CMO report PDF')
    parser.add_argument('--cmo-name', help='Name of CMO (e.g., "COSON Nigeria")')
    parser.add_argument('--output-dir', help='Output directory', default='./outputs')
    parser.add_argument('--no-db', action='store_true', help='Skip database ingestion')
    
    args = parser.parse_args()
    
    # Initialize pipeline
    pipeline = OrderSoundsPipeline(config={
        'output_dir': args.output_dir
    })
    
    # Process report
    result = pipeline.process_cmo_report(
        pdf_path=args.pdf_path,
        cmo_name=args.cmo_name,
        write_db=not args.no_db
    )
    
    # Exit with status code
    sys.exit(0 if result.get('status') != 'failed' else 1)


if __name__ == '__main__':
    main()
